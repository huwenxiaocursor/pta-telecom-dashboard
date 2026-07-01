#!/usr/bin/env python3
"""
Fetches Pakistan macro-economic data (policy rate, forex reserves, PKR/USD
rate, remittances, CPI) from SBP/PBS official sources, merges into a
permanent history file, and injects a rolling window into macro_index.html.

Sources:
- SBP "war-current.asp" snapshot page: policy rate, forex reserves, FX rate
  (all three are on this one static HTML page — no PDF/LLM needed)
- SBP Homeremit_Arch.xlsx: workers' remittances (updates infrequently)
- PBS Monthly Review PDF: CPI (needs PyMuPDF text extraction + DeepSeek
  structured-extraction fallback, since PBS only publishes prose/PDF, no
  structured table)

Trade data is intentionally NOT automated in this version — PBS's trade
data URL/format was not confirmed reliable; Section③ 贸易 stays manual.
GDP/fiscal/sector data stays manual (Economic Survey, published ~once/year);
check_new_economic_survey() only alerts when a newer survey is detected.
"""

import datetime
import io
import json
import os
import pathlib
import re
import shutil
import urllib.request

BASE_DIR       = pathlib.Path(__file__).resolve().parent
HTML_FILE      = BASE_DIR.parent / "macro_index.html"
LOG_FILE       = BASE_DIR / "macro_update_log.txt"
HISTORY_FILE   = BASE_DIR / "macro_history.json"
FY_STATE_FILE  = BASE_DIR / "macro_known_fy.json"
GDP_ALERT_FILE = BASE_DIR / "macro_gdp_update_needed.txt"

DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")

WAR_CURRENT_URL   = "https://www.sbp.org.pk/ecodata/rates/war/war-current.asp"
REMIT_XLSX_URL    = "https://www.sbp.org.pk/assets/document/Homeremit_Arch.xlsx"
PBS_PRICE_STATS_URL = "https://www.pbs.gov.pk/price-statistics/"
SURVEY_LIST_URL   = "https://www.finance.gov.pk/survey_archieve.html"

HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}

MONTH_ABBR = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}
MONTH_FULL = {m: i + 1 for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June",
     "July", "August", "September", "October", "November", "December"])}


def log(msg):
    ts = datetime.datetime.now().isoformat(timespec="seconds")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def fetch(url, timeout=30):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        charset = r.headers.get_content_charset() or "utf-8"
        return r.read().decode(charset, errors="ignore")


def fetch_binary(url, timeout=30):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def load_history():
    if HISTORY_FILE.exists():
        return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    return {}


def save_history(history):
    HISTORY_FILE.write_text(
        json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def merge_record_into_history(history, key, period, fields):
    """Add a single dated record into history[key][period]; never overwrite
    a field that's already present (mirrors merge_monthly_into_history in
    update_pta_dashboard.py, generalized to non-monthly periods)."""
    if key not in history:
        history[key] = {}
    if period not in history[key]:
        history[key][period] = {}
    added = []
    for k, v in fields.items():
        if k not in history[key][period]:
            history[key][period][k] = v
            added.append(k)
    return added


def latest_period(history, key):
    periods = sorted(history.get(key, {}).keys())
    return periods[-1] if periods else None


def recent_periods(history, key, n):
    return sorted(history.get(key, {}).keys())[-n:]


def parse_sbp_date(s):
    """Parses SBP's inconsistent date formats: '19- June - 2026', '30 - Jun - 2026'."""
    m = re.search(r"(\d{1,2})\s*-?\s*([A-Za-z]+)\s*-?\s*(\d{4})", s)
    if not m:
        return None
    day, mon, year = m.groups()
    mon_num = MONTH_ABBR.get(mon[:3].title()) or MONTH_FULL.get(mon.title())
    if not mon_num:
        return None
    return f"{int(year):04d}-{mon_num:02d}-{int(day):02d}"


def validate_bounds(value, lo, hi):
    try:
        return lo <= float(value) <= hi
    except (TypeError, ValueError):
        return False


# ─── DeepSeek structured extraction ────────────────────────────────────────

def extract_json_via_deepseek(system_prompt, user_text):
    """Sends text to DeepSeek and expects a strict JSON object back.
    Returns None on any failure (no key, HTTP error, non-JSON) — caller
    must treat None as 'skip this metric this run', never as zero/blank."""
    if not DEEPSEEK_API_KEY:
        log("  DeepSeek skipped: DEEPSEEK_API_KEY not set")
        return None

    payload = json.dumps({
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_text},
        ],
        "response_format": {"type": "json_object"},
        "max_tokens": 600,
        "temperature": 0,
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.deepseek.com/chat/completions",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        content = result["choices"][0]["message"]["content"]
        return json.loads(content)
    except Exception as e:
        log(f"  DeepSeek extraction error: {e}")
        return None


# ─── Fetcher 1/2/3: war-current.asp (policy rate, reserves, FX rate) ───────

def fetch_war_current_page():
    return fetch(WAR_CURRENT_URL)


def fetch_policy_rate(html, history):
    m_rate = re.search(r"SBP Policy Rate</p>\s*<span[^>]*>([\d.]+)%", html)
    m_ceil = re.search(r"Ceiling\)?\s*Rate</p>\s*<span[^>]*>([\d.]+)%", html)
    m_floor = re.search(r"\(Floor\)\s*Rate</p>\s*<span[^>]*>([\d.]+)%", html)
    if not m_rate:
        raise RuntimeError("未找到 SBP Policy Rate，页面结构可能已变化")

    rate = float(m_rate.group(1))
    if not validate_bounds(rate, 0, 30):
        raise RuntimeError(f"政策利率数值超出合理范围：{rate}")

    prev_period = latest_period(history, "policyRate")
    prev_rate = history.get("policyRate", {}).get(prev_period, {}).get("rate") if prev_period else None

    if prev_rate is not None and abs(rate - prev_rate) < 1e-9:
        return None  # unchanged since last recorded event, nothing new to add

    today = datetime.date.today().isoformat()
    change_bp = round((rate - prev_rate) * 100) if prev_rate is not None else 0
    fields = {"rate": rate, "changeBp": change_bp}
    if m_ceil:
        fields["corridorUpper"] = float(m_ceil.group(1))
    if m_floor:
        fields["corridorLower"] = float(m_floor.group(1))
    fields["note"] = "自动检测到利率变动" if prev_rate is not None else "首次记录"
    return today, fields


def fetch_reserves(html):
    anchor = html.find("Liquid Foreign Exchange Reserves")
    if anchor == -1:
        raise RuntimeError("未找到外汇储备区块，页面结构可能已变化")
    window = html[anchor:anchor + 2000]

    date_m = re.search(r"As on ([^<]+)</h5>", window)
    if not date_m:
        raise RuntimeError("未找到外汇储备日期")
    date_iso = parse_sbp_date(date_m.group(1))
    if not date_iso:
        raise RuntimeError(f"外汇储备日期解析失败：{date_m.group(1)!r}")

    pairs = re.findall(r'<p class="text-white mb-0">([^<]+)</p>\s*<h4[^>]*>\s*([\d,\.]+)\s*</h4>', window)
    fields = {}
    for label, val in pairs:
        num = float(val.replace(",", ""))
        if "SBP" in label:
            fields["sbp"] = num
        elif "Bank" in label:
            fields["banks"] = num
        elif "Total" in label:
            fields["total"] = num

    if "sbp" not in fields:
        raise RuntimeError("未能解析出 SBP 持有储备数值")
    if not validate_bounds(fields["sbp"], 0, 200000):
        raise RuntimeError(f"外汇储备数值超出合理范围：{fields['sbp']}")

    return date_iso, fields


def fetch_fx_rate(html):
    anchor = html.find("USD/ PKR Rates")
    if anchor == -1:
        anchor = html.find("USD/PKR Rates")
    if anchor == -1:
        raise RuntimeError("未找到汇率区块，页面结构可能已变化")
    window = html[anchor:anchor + 2000]

    date_m = re.search(r"As on ([^<]+)</p>", window)
    m2m_m = re.search(r"M2M Revaluation Rate</p>\s*<h4[^>]*>\s*([\d.]+)", window, re.S)
    bid_m = re.search(r"BID</p>\s*<h4[^>]*>\s*([\d.]+)", window, re.S)
    offer_m = re.search(r"Offer</p>\s*<h4[^>]*>\s*([\d.]+)", window, re.S)

    if not date_m or not bid_m or not offer_m:
        raise RuntimeError("未能解析出汇率数值，页面结构可能已变化")

    date_iso = parse_sbp_date(date_m.group(1))
    if not date_iso:
        raise RuntimeError(f"汇率日期解析失败：{date_m.group(1)!r}")

    bid, offer = float(bid_m.group(1)), float(offer_m.group(1))
    if not (validate_bounds(bid, 100, 500) and validate_bounds(offer, 100, 500)):
        raise RuntimeError(f"汇率数值超出合理范围：bid={bid}, offer={offer}")

    fields = {"bid": bid, "offer": offer, "mid": round((bid + offer) / 2, 2)}
    if m2m_m:
        fields["m2m"] = float(m2m_m.group(1))
    return date_iso, fields


# ─── Fetcher 4: remittances (SBP Excel archive) ────────────────────────────

def fetch_remittances():
    import openpyxl

    data = fetch_binary(REMIT_XLSX_URL)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "onwards" in s.lower()), None)
    if not sheet_name:
        raise RuntimeError(f"未找到包含'onwards'的工作表，实际工作表：{wb.sheetnames}")
    ws = wb[sheet_name]

    HEADER_ROW, CASH_ROW = 5, 6
    COUNTRY_ROWS = {}  # row -> our JSON key
    for r in range(7, 12):
        name = (ws.cell(row=r, column=3).value or "").strip()
        if name == "Saudi Arabia":
            COUNTRY_ROWS[r] = "Saudi Arabia"
        elif name.replace(".", "") == "UAE" or name == "U.A.E.":
            COUNTRY_ROWS[r] = "UAE"
        elif name.replace(".", "") == "UK" or name == "U.K.":
            COUNTRY_ROWS[r] = "UK"

    last_col = None
    for c in range(ws.max_column, 3, -1):
        if ws.cell(row=HEADER_ROW, column=c).value is not None:
            last_col = c
            break
    if last_col is None:
        raise RuntimeError("未能在侨汇表中找到任何月份列")

    def month_iso(val):
        if isinstance(val, datetime.datetime):
            return f"{val.year:04d}-{val.month:02d}"
        m = re.match(r"([A-Za-z]{3})-(\d{2})", str(val).strip())
        if not m:
            return None
        mon, yy = m.groups()
        mon_num = MONTH_ABBR.get(mon.title())
        if not mon_num:
            return None
        return f"{2000 + int(yy):04d}-{mon_num:02d}"

    iso = month_iso(ws.cell(row=HEADER_ROW, column=last_col).value)
    if not iso:
        raise RuntimeError(f"月份解析失败：{ws.cell(row=HEADER_ROW, column=last_col).value!r}")

    total = ws.cell(row=CASH_ROW, column=last_col).value
    if total is None:
        raise RuntimeError(f"{iso} 列没有总额数据")
    total_usd = float(total) * 1_000_000

    fields = {"totalUsd": total_usd}

    prev_total = ws.cell(row=CASH_ROW, column=last_col - 1).value
    if prev_total:
        fields["momPct"] = round((total / float(prev_total) - 1) * 100, 1)
    if last_col - 12 >= 4:
        yoy_total = ws.cell(row=CASH_ROW, column=last_col - 12).value
        if yoy_total:
            fields["yoyPct"] = round((total / float(yoy_total) - 1) * 100, 1)

    top = {}
    for r, key in COUNTRY_ROWS.items():
        v = ws.cell(row=r, column=last_col).value
        if v is not None:
            top[key] = float(v) * 1_000_000
    if top:
        fields["topCountries"] = top

    return iso, fields


def compute_fy_cumulative(history, fy_label):
    """FY cumulative = frozen baseline (cumulative through baseline['throughMonth'])
    + sum of monthly entries strictly AFTER throughMonth. Only months after the
    baseline cutoff are summed, so backfilled months at/before the cutoff (e.g. a
    remittances fetch that lags and returns an older month) are never double-counted
    against the baseline that already implicitly includes them."""
    baseline = history.get("remittancesFYBaseline", {}).get(fy_label)
    if not baseline:
        return None
    through_month = baseline["throughMonth"]
    total = baseline["cumulativeUsd"]
    for period, rec in history.get("remittances", {}).items():
        if period > through_month and "totalUsd" in rec:
            total += rec["totalUsd"]
    return total


# ─── Fetcher 5: CPI (PBS Monthly Review PDF, PyMuPDF + DeepSeek) ───────────

def find_latest_cpi_pdf_url():
    html = fetch(PBS_PRICE_STATS_URL)
    m = re.search(r'href="(https://www\.pbs\.gov\.pk/[^"]*Monthly-Review-[^"]*\.pdf)"', html)
    if not m:
        raise RuntimeError("未在PBS价格统计页找到 Monthly Review PDF 链接")
    url = m.group(1)
    name_m = re.search(r"Monthly-Review-([A-Za-z]+)-(\d{4})\.pdf", url)
    if not name_m:
        raise RuntimeError(f"无法从文件名解析月份：{url}")
    mon, year = name_m.groups()
    mon_num = MONTH_FULL.get(mon.title())
    if not mon_num:
        raise RuntimeError(f"无法识别月份名：{mon}")
    return url, f"{int(year):04d}-{mon_num:02d}"


def extract_cpi_brief_text(pdf_bytes):
    import fitz
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = ""
    for page in doc[:3]:
        text += page.get_text()
    idx = text.find("Inflation in")
    if idx == -1:
        raise RuntimeError("PDF中未找到 'Inflation in Brief' 段落")
    section = text[idx:idx + 2200]
    section = re.sub(r"[ \t]+", " ", section)
    # Recover the common "leading digit separated by whitespace" artifact,
    # e.g. "1 1.1 %" -> "11.1%" (confirmed pattern in real PBS PDF extracts)
    section = re.sub(r"(?<!\d)(\d)\s+(\d\s*\.\s*\d)\s*%", lambda mo: mo.group(1) + mo.group(2).replace(" ", "") + "%", section)
    section = re.sub(r"\s+", " ", section)
    return section.strip()


def fetch_cpi(history):
    url, month_iso = find_latest_cpi_pdf_url()
    if month_iso in history.get("cpi", {}) and "yoy" in history["cpi"][month_iso]:
        return None  # already have this month, nothing new
    pdf_bytes = fetch_binary(url, timeout=60)
    brief = extract_cpi_brief_text(pdf_bytes)

    ref_month = latest_period(history, "cpi")
    ref = history.get("cpi", {}).get(ref_month, {}) if ref_month else {}

    system_prompt = (
        "你是数据抽取助手。以下文本来自巴基斯坦统计局(PBS)月度CPI报告的'Inflation in Brief'段落，"
        "该PDF文本提取工具有一个已知缺陷：10-19之间的数字有时会丢失开头的'1'"
        "（有时表现为'1 1.1'这种被空格断开的形式，有时完全丢失变成'1.1'）。"
        "请仔细通读全文语义（每句话都明确说明了这个数字对应'本月/上月/去年同月'哪个含义），"
        "结合下方提供的'已知上月官方数据'核对文本中'上月'对应的数字是否吻合——"
        "如果不吻合（比如文本显示1.8%而已知数据是11.8%），说明该数字丢了开头的'1'，"
        "那么本月同一句子里的数字也可能有同样问题，请按语义合理性判断是否需要补回'1'。"
        "只输出本月（最新月份）的数据，严格按以下JSON schema输出，数值单位为百分比，"
        "找不到的字段设为null：\n"
        '{"yoy": 全国CPI同比, "mom": 全国CPI环比, "urbanYoy": 城市CPI同比, '
        '"ruralYoy": 农村CPI同比, "spiYoy": SPI同比, "nfneUrbanYoy": null, "nfneRuralYoy": null}'
    )
    user_text = (
        f"已知上月({ref_month})官方数据供比对：{json.dumps(ref, ensure_ascii=False)}\n\n"
        f"本月报告原文：\n{brief}"
    )
    result = extract_json_via_deepseek(system_prompt, user_text)
    if not result:
        return None

    if not validate_bounds(result.get("yoy"), 0, 50):
        log(f"  CPI抽取结果超出合理范围，丢弃：{result}")
        return None
    if ref.get("yoy") is not None and abs(result["yoy"] - ref["yoy"]) > 8:
        log(f"  CPI同比与上月({ref.get('yoy')})差距过大（{result['yoy']}），怀疑抽取有误，丢弃")
        return None

    fields = {k: v for k, v in result.items() if v is not None}
    return month_iso, fields


# ─── Fetcher 6: trade — not implemented in v1 ──────────────────────────────

def fetch_trade():
    return None


# ─── Annual Economic Survey alert (mirrors check_new_qos_reports) ──────────

def check_new_economic_survey():
    try:
        html = fetch(SURVEY_LIST_URL)
    except Exception as e:
        log(f"  Economic Survey 页面抓取失败（不影响其他更新）：{e}")
        return

    years = re.findall(r"Economic Survey (20\d\d-\d\d)", html)
    if not years:
        return
    latest = sorted(years)[-1]

    known = {}
    if FY_STATE_FILE.exists():
        known = json.loads(FY_STATE_FILE.read_text(encoding="utf-8"))
    last_known = known.get("last_known_survey_fy")

    if latest != last_known:
        GDP_ALERT_FILE.write_text(
            f"检测到新的 Pakistan Economic Survey {latest}，请人工核对并更新 macro_index.html "
            f"中标注为 MANUAL 的 GDP/财政/产业结构/贸易板块。\n来源页面：{SURVEY_LIST_URL}\n",
            encoding="utf-8",
        )
        known["last_known_survey_fy"] = latest
        FY_STATE_FILE.write_text(json.dumps(known, ensure_ascii=False, indent=2), encoding="utf-8")
        log(f"发现新的 Economic Survey {latest}，已写入 macro_gdp_update_needed.txt 提醒人工处理")
    elif GDP_ALERT_FILE.exists():
        pass  # leave existing alert until a human clears it


# ─── HTML injection ─────────────────────────────────────────────────────────

def build_macro_data_js(history):
    def jl(v):
        return json.dumps(v, ensure_ascii=False)

    today = datetime.date.today().isoformat()
    data = {"asOf": today, "generatedAt": today}

    def with_prev(key, latest_p, extra_n=1):
        periods = sorted(history.get(key, {}).keys())
        idx = periods.index(latest_p)
        if idx - extra_n >= 0:
            prev_p = periods[idx - extra_n]
            return {"period": prev_p, **history[key][prev_p]}
        return None

    fx_p = latest_period(history, "fxRate")
    if fx_p:
        data["fx"] = {"date": fx_p, **history["fxRate"][fx_p], "prev": with_prev("fxRate", fx_p)}
        pts = recent_periods(history, "fxRate", 8)
        data["chart"] = data.get("chart", {})
        data["chart"]["fxLabels"] = pts
        data["chart"]["fxVals"] = [history["fxRate"][p].get("mid") for p in pts]

    res_p = latest_period(history, "reserves")
    if res_p:
        data["reserves"] = {"date": res_p, **history["reserves"][res_p], "prev": with_prev("reserves", res_p)}
        pts = recent_periods(history, "reserves", 5)
        data["chart"] = data.get("chart", {})
        data["chart"]["reservesLabels"] = pts
        data["chart"]["reservesVals"] = [history["reserves"][p].get("sbp") for p in pts]

    cpi_p = latest_period(history, "cpi")
    if cpi_p:
        data["cpi"] = {"month": cpi_p, **history["cpi"][cpi_p], "prev": with_prev("cpi", cpi_p)}
        pts = recent_periods(history, "cpi", 8)
        data["chart"] = data.get("chart", {})
        data["chart"]["cpiLabels"] = pts
        data["chart"]["cpiVals"] = [history["cpi"][p].get("yoy") for p in pts]

    pr_p = latest_period(history, "policyRate")
    if pr_p:
        data["policyRate"] = {"date": pr_p, **history["policyRate"][pr_p]}
        pts = recent_periods(history, "policyRate", 10)
        data["chart"] = data.get("chart", {})
        data["chart"]["policyRateLabels"] = pts
        data["chart"]["policyRateVals"] = [history["policyRate"][p].get("rate") for p in pts]

    remit_p = latest_period(history, "remittances")
    if remit_p:
        data["remittances"] = {"month": remit_p, **history["remittances"][remit_p]}
        data["remittances"]["fy26CumulativeUsd"] = compute_fy_cumulative(history, "FY26")

    return f"const MACRO_DATA = {jl(data)};"


def update_html(macro_data_js):
    html = HTML_FILE.read_text(encoding="utf-8")
    START = "// ===AUTO-MACRO-DATA-START==="
    END = "// ===AUTO-MACRO-DATA-END==="
    new_block = f"{START}\n{macro_data_js}\n{END}"

    existing = re.search(re.escape(START) + r".*?" + re.escape(END), html, re.DOTALL)
    if existing:
        html = html.replace(existing.group(0), new_block)
    else:
        anchor = html.find("<script>\nconst greenA")
        if anchor == -1:
            raise RuntimeError("HTML中找不到脚本插入点，需人工检查（首次运行需要先手动完成 macro_index.html 改造）")
        insert_at = html.find("\n", anchor) + 1
        html = html[:insert_at] + new_block + "\n\n" + html[insert_at:]

    backup = None
    if not os.environ.get("CI"):
        backup = HTML_FILE.with_suffix(HTML_FILE.suffix + f".bak-{datetime.date.today().isoformat()}")
        shutil.copy2(HTML_FILE, backup)
    HTML_FILE.write_text(html, encoding="utf-8")
    return backup


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    log("=" * 50)
    log("Macro update started")

    history = load_history()
    updated, skipped = [], []

    try:
        war_html = fetch_war_current_page()
    except Exception as e:
        war_html = None
        log(f"警告：war-current.asp 抓取失败，本次跳过利率/储备/汇率三项：{e}")
        skipped += ["policyRate", "reserves", "fxRate"]

    if war_html is not None:
        try:
            result = fetch_policy_rate(war_html, history)
            if result:
                period, fields = result
                merge_record_into_history(history, "policyRate", period, fields)
                updated.append(f"policyRate({period}={fields['rate']}%)")
                log(f"政策利率更新：{period} = {fields['rate']}%")
            else:
                log("政策利率：与已记录数值一致，无需新增")
        except Exception as e:
            log(f"警告：政策利率抓取失败：{e}")
            skipped.append("policyRate")

        try:
            period, fields = fetch_reserves(war_html)
            added = merge_record_into_history(history, "reserves", period, fields)
            if added:
                updated.append(f"reserves({period})")
                log(f"外汇储备更新：{period} = {fields}")
            else:
                log(f"外汇储备：{period} 已存在，无需新增")
        except Exception as e:
            log(f"警告：外汇储备抓取失败：{e}")
            skipped.append("reserves")

        try:
            period, fields = fetch_fx_rate(war_html)
            added = merge_record_into_history(history, "fxRate", period, fields)
            if added:
                updated.append(f"fxRate({period})")
                log(f"汇率更新：{period} = {fields}")
            else:
                log(f"汇率：{period} 已存在，无需新增")
        except Exception as e:
            log(f"警告：汇率抓取失败：{e}")
            skipped.append("fxRate")

    try:
        result = fetch_remittances()
        if result:
            period, fields = result
            added = merge_record_into_history(history, "remittances", period, fields)
            if added:
                updated.append(f"remittances({period})")
                log(f"侨汇更新：{period} = {fields}")
            else:
                log(f"侨汇：{period} 已存在，无需新增")
    except Exception as e:
        log(f"警告：侨汇抓取失败：{e}")
        skipped.append("remittances")

    try:
        result = fetch_cpi(history)
        if result:
            period, fields = result
            merge_record_into_history(history, "cpi", period, fields)
            updated.append(f"cpi({period})")
            log(f"CPI更新：{period} = {fields}")
        else:
            log("CPI：无新月份或抽取被丢弃，跳过")
    except Exception as e:
        log(f"警告：CPI抓取失败：{e}")
        skipped.append("cpi")

    fetch_trade()  # not implemented in v1, always returns None

    save_history(history)

    macro_data_js = build_macro_data_js(history)
    try:
        backup = update_html(macro_data_js)
        backup_note = f"（已备份至 {backup.name}）" if backup else "（CI环境，跳过本地备份）"
        log(f"HTML写入成功{backup_note}")
    except Exception as e:
        log(f"错误：HTML写入失败：{e}")

    try:
        check_new_economic_survey()
    except Exception as e:
        log(f"警告：Economic Survey 检查失败（不影响其他更新）：{e}")

    log(f"本次更新：{updated if updated else '无'}")
    log(f"本次跳过：{skipped if skipped else '无'}")
    log("Macro update complete")
    log("=" * 50)


if __name__ == "__main__":
    main()
